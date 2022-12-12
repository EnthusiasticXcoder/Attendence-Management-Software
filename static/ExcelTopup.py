import tkinter as tk
import customtkinter as ctk
from tkinter import ttk
from threading import Thread
from tkinter import messagebox
from PIL import Image

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from xlcalculator import ModelCompiler
from xlcalculator import Evaluator

try:
    from utilits import Drive_uplode
    from static.ScrollFrame import ScrollFrame
except: pass


class Cell(ctk.CTkLabel):
    def __init__(self, *args, 
                bg_color='transparent', 
                fg_color=None, 
                text_color=None, 
                corner_radius=None, 
                width=140, 
                height=28, 
                text=" ", 
                font=None, 
                anchor="nw", 
                **kwargs):

        super().__init__(*args, 
                            bg_color=bg_color, 
                            fg_color=fg_color, 
                            text_color=text_color, 
                            corner_radius=0,
                            width=int(len(str(text))*9.1), 
                            height=40, 
                            text=text, 
                            font=("Roboto Medium", -16,'bold'), 
                            anchor=anchor, **kwargs)

class Tabularcell(ctk.CTkEntry):
    def __init__(self, *args, text,column,maxr=False,
                bg_color='transparent', 
                fg_color=None, 
                text_color=None, 
                placeholder_text_color=None, 
                font=('Times New Roman', 14, 'bold'), 
                placeholder_text=None, 
                corner_radius=0, 
                border_width=2, 
                border_color=None, 
                width=30, 
                height=28, 
                state=tk.DISABLED, 
                textvariable: tk.Variable = None, **kwargs):

        width= 200 if column==2 else 30
        width= 50 if column in [22,31,32] else width
        width= 150 if column==1 else width 
        width= 34 if column==0 else width
        width=384 if maxr==True and column==0 else width

        textvariable=tk.StringVar(value=text)
        justify= 'left' if width==200 else 'center'

        super().__init__(*args, bg_color=bg_color, 
                        fg_color=fg_color, 
                        text_color=text_color, 
                        placeholder_text_color=placeholder_text_color, 
                        font=font, 
                        placeholder_text=placeholder_text, 
                        corner_radius=corner_radius, 
                        border_width=border_width, 
                        border_color=border_color, 
                        width=width, 
                        height=height, 
                        state=state, 
                        justify=justify,
                        textvariable=textvariable, **kwargs)


class Headcell(ctk.CTkTextbox):
    def __init__(self, *args, text,column,head=False,
                bg_color='transparent', 
                fg_color=None, 
                border_color=None, 
                border_width=10, 
                corner_radius=2, 
                text_font=('Times New Roman', 11, 'bold'), 
                text_color=None, 
                width=200, height=80, **kwargs):
        
        border_color=("#565B5E",'#565B5E')
        if column<2 : text_font=('Times New Roman', 13, 'bold')
        if column==2 : text_font=('Times New Roman', 17, 'bold')
        if head:
            width= 0
        else:
            width= 200 if column==2 else 30
            width= 50 if column in [22,31,32] else width
            width= 150 if column==1 else width 
            width= 34 if column==0 else width

        height= height-25 if 2<column<22 else height
        height= height-25 if 22<column<31 else height

        height= 25 if head==True else height

        super().__init__(*args, bg_color=bg_color, 
                        fg_color=fg_color, 
                        border_color=border_color, 
                        border_width=0, 
                        corner_radius=corner_radius, 
                        font=text_font, 
                        text_color=text_color, 
                        width=width, height=height,
                        activate_scrollbars=False,border_spacing=0, **kwargs)

        self.insert('0.0',f"{text}")
        self.configure(state="disabled")


def Display(frame,ws,evaluator):
    MAXC=ws.max_column
    MAXR=0
    while True:
        if ws['A'+str(MAXR+10)].value == None  :
            break  
        MAXR+=1
    MAXR +=11

    frame1=ctk.CTkFrame(master=frame,width=1200)
    frame1.grid(row=0,column=0,sticky='nsew')

    indexc=[i for i in range(MAXC)]
    indexr=[i for i in range(7)]

    frame1.rowconfigure(indexr,weight=0)
    frame1.columnconfigure(indexc,weight=1)

    frame2=ctk.CTkFrame(master=frame)
    frame2.grid(row=1,column=0,sticky='nsew')

    for row in range(1,MAXR+1):
        if row in [7,8,9] : continue
        span=1
        values=[]
        merge=[]
        col=[]
            
        for column,i in enumerate(ws[str(row)]):
            if 'MergedCell' in str(i) :
                span+=1
            else:
                if column==0:
                    if row==9:
                        try:
                            arr=i.value.split("\n")
                            value=f'{arr[0]}\n{arr[1]}\n{arr[2]}'
                        except :
                            pass
                    else:
                        values.append(i.value)
                    col.append(column)
                else:
                    if column in [22,31,32] and 9<row<MAXR or row==MAXR :
                        if i.value==None:
                            values.append(i.value)
                        else:
                            alfacol=get_column_letter(column+1)
                            val = evaluator.evaluate(f'{ws.title}!{alfacol}{row}')
                            values.append(val) if val != None else values.append(0)
                    else:
                        values.append(i.value)
                    merge.append(span)
                    col.append(column)
                span=1
        merge.append(span)
    
        s= 'n' if row in[1,2,'THEORY','PRACTICAL'] else 'nw'

        table= True if row>9 else False

        if [None]*MAXC==values : 
            if table :
                row-=2
                frame2.rowconfigure(row,minsize=55)
            else: frame1.rowconfigure(row,minsize=40)
            continue

        if table:
            rowframe=ctk.CTkFrame(master=frame2,corner_radius=0)
            rowframe.grid(row=row-1,column=0,sticky='nsew')
            for value,span,column in zip(values,merge,col):
                if value == None : continue
                if value==0 and row==MAXR: continue
                maxr= True if row==MAXR else False 
                Tabularcell(master=rowframe,text=value,column=column,maxr=maxr).grid(row=0,column=column,columnspan=span,sticky=s)
        elif row<7:
            for value,span,column in zip(values,merge,col):
                if value==None : 
                    continue
                else:
                    Cell(master=frame1,text=value,corner_radius=0).grid(row=row-1,column=column,columnspan=span,sticky=s)    

    
    headframe=ctk.CTkFrame(master=frame2,corner_radius=0)
    headframe.grid(row=0,column=0,sticky='nsew')

    headframe.rowconfigure(0,weight=1)
    headframe.rowconfigure(1,weight=2)

    for column,i in enumerate(ws['8']) :
        val= " " if i.value==None else i.value
        if val in ['THEORY',"PRACTICAL"] : 
            headwidth=0
            headcell=Headcell(master=headframe,column=column,text=val,height=25,head=True)
            headcell.grid(row=0,column=column,columnspan=19,sticky='nw')
        row,span=0,1
        if column>2 and column not in [22,31,32] : row+=1
        else : span=2 

        if 2<column<22 or 22<column<31 :
            try:
                char=get_column_letter(column+1)
                val=ws[f'{char}9'].value
                if val==None : 
                    continue
                arr=val.split("\n")
                val=f'{arr[0]}\n{arr[1]}\n{arr[2]}'
                headwidth+=1
                headcell.configure(width=int(headwidth*30.42))
            except :
                continue
        Headcell(master=headframe,column=column,text=val).grid(row=row,column=column,rowspan=span,sticky='nw')

def See_Excel(WBname):
    Top=ctk.CTkToplevel()
    Top.geometry('1000x500')
    if '/' in WBname:
        folder=WBname.split("/")[-2]
        wbtitle=WBname.split("/")[-1]
    elif '\\' in WBname:
        folder=WBname.split("\\")[-2]
        wbtitle=WBname.split("\\")[-1]

    Top.title(f" SHEET  VIEW        \t\t\t\t\t\t\t\t\t\t\t{wbtitle}")
    DownloadFrame=ctk.CTkFrame(master=Top,fg_color=('#EBEBEC','#212325'))
    DownloadFrame.pack(pady=20,fill=tk.X)

    img = ctk.CTkImage(light_image=Image.open("images/download.png"),dark_image=Image.open("images/download.png"),size=(35,35))
    B_Download = ctk.CTkButton(master=DownloadFrame, 
                                text = "Download",
                                fg_color="green",text_color="white",image=img,
                                font=("times new roman",15,'bold'),
                                command= lambda : Thread(target=Drive_uplode.download_files , args=( folder , wbtitle , True)).start())  
    B_Download.pack(side=tk.RIGHT,padx=100,ipadx=50)

    # Create A Main Frame
    main_frame = ctk.CTkFrame(Top)
    main_frame.pack(fill=tk.BOTH, expand=1)

    Scroll=ScrollFrame(master=main_frame,PlaceScrollBar='Both',binding_root=Top)
    Scroll.pack(fill=tk.BOTH,expand=1)
    Scroll.updateScrollBar()

    my_notebook = ttk.Notebook(Scroll.Frame)
    my_notebook.pack(padx=10)

    style=ttk.Style()
    #style
    noteback= '#2A2D2E' if ctk.get_appearance_mode()=='Dark' else "#D1D5D8"
    tabbg= "#2B2B2B" if ctk.get_appearance_mode()=='Dark' else '#D1D5D8' 
    tabfg= "white" if ctk.get_appearance_mode()=='Dark' else 'black' 
    style.theme_use('clam')
    style.configure("TNotebook",tabmargins=[10, 1, 5 , 0],background=noteback)
    style.configure("TNotebook.Tab",padding= [1, 5] , font=('consolas italic', 14), borderwidth=[5])
    style.map("TNotebook.Tab",foreground=[("selected", tabfg), ('!active', tabfg), ('active', tabfg)],
                                 background=[("selected", "#2A73AC"), ('!active', tabbg ), ('active', "#144870")],
                                font=[("active", ('consolas italic', 14,'bold')) ,("!active", ('consolas italic', 15)),("selected", ('consolas italic', 14,'bold'))],
                                expand= [("selected", [1, 1, 1, 0])])
    style.layout("Tab",
                        [('Notebook.tab', {'sticky': 'nswe', 'children':
                            [('Notebook.padding', {'side': 'top', 'sticky': 'nswe', 'children':
                                [('Notebook.label', {'side': 'top', 'sticky': ''})],
                            })],
                        })]
                    ) 

    try :
        wb=load_workbook(WBname)
    except :
        messagebox.showerror("Error","Error! Unable To Open\nPlease Try Again")

    compiler = ModelCompiler()
    new_model = compiler.read_and_parse_archive(WBname)
    evaluator = Evaluator(new_model)

    for i in wb.sheetnames :
        ws=wb[i]
        frame=ctk.CTkFrame(master=my_notebook,
                            height=ws.max_row*30,
                            width=ws.max_column*43,
                            )
        frame.pack(fill=tk.BOTH, expand=1)
        my_notebook.add(frame, text=i)
        Display(frame,ws,evaluator)


if __name__=="__main__":
    
    r=ctk.CTk()
    r.title("CustomTkinter complex_example.py")
    r.state('zoomed')

    ctk.set_appearance_mode("light")  # Modes: "System" (standard), "Dark", "Light"
    ctk.set_default_color_theme("blue")
    # ============ create two frames ============
    ctk.CTkButton(master=r,command=lambda :See_Excel("workbooks/Attendence.xlsx")).pack()
    ModeChange = ctk.CTkOptionMenu(master=r,
                                    values=["Light", "Dark", "System"],
                                    command=ctk.set_appearance_mode)
    ModeChange.pack()

    r.mainloop()
