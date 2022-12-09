from threading import Thread
from tkinter import filedialog,messagebox
import shutil
import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from utilits import Drive_uplode


def import_(user,Function):
    fileto=filedialog.askopenfilename(
        initialdir='D:/',
        title="Import A File",
        filetypes=[("xlsx files","*.xlsx")]
        )
    if fileto:
        try:
            fileto=r"{}".format(fileto)
            fil=os.getcwd()
            f=fileto.split("/")
            f=f[-1]
            filefr=os.path.join(fil,f"workbooks\\{user}\\{f}")
            filefr=r"{}".format(filefr)
            shutil.copyfile(fileto,filefr)
            
            if Drive_uplode.DStoreID!=None: Thread(target=Drive_uplode.DStoreID.Update,args=(user,)).start()
            Thread(target=Drive_uplode.upload_files,args=(user,f,Function)).start()
            messagebox.showinfo("","File imported Sucessfully")
        except :
            messagebox.showerror("Error","Cannot Import File")

def createnewws(user,Excel,title):
    try:
        wb=load_workbook(f"workbooks\{user}\{Excel}")
        ws1=wb["Sheet1"]
        ws=wb.copy_worksheet(ws1)
        ws.title= title

        with open("users/sheets.txt","r") as f:
            data = f.readlines()
        
        with open("users/sheets.txt","w") as f:
            for names in data :
                if user in names and Excel in names:
                    continue
                f.write(names)
            w=f'{user} {Excel} {title} \n'
            f.write(w)

        wb.save(f"workbooks\{user}\{Excel}")
        Thread(target=Drive_uplode.upload_files , args=(user,Excel)).start()
        Thread(target=Drive_uplode.upload_files , args=('user','sheets.txt')).start()
        messagebox.showinfo("Sucess","New sheet created")
    except:
        messagebox.showerror("Error","Unable to create worksheet")
        return

def numberofentry(ws):
    n=0
    while True:
        if ws['A'+str(n+10)].value == None  :
            break
        n=n+1
    return n

def cheackplace(ws):
    for row in range(4,23):
        char = get_column_letter(row)
        if ws[char+'9'].value == None  :
            break
    return char

def entertheory(backfun,object,user,Excel,date,time,roll_list):
    try:
        try:
            wb=load_workbook(f"workbooks\\{user}\\{Excel}")
        except:
            messagebox.showerror("Error","Workbook not found")
            return
        
        if len(wb.sheetnames)==1:
            messagebox.showerror("","Create new sheet")
            return
            
        with open("users/sheets.txt","r") as f :
            for name in f.readlines():
                if user+" "+Excel in name:
                    b=name.split(" ")
                    sheet=b[-2]
                    ws=wb[sheet]
                    break
            else:
                with open("users/sheets.txt","a") as f1 :
                    ws=wb.active
                    w=user+" "+Excel+" "+ws.title+" \n"
                    f1.write(w)
        char=cheackplace(ws)
        n=numberofentry(ws)
        if char=="V":
            messagebox.showerror("","Sheet is full! Create new sheet And Try Again")
            return
        object.extend_start()
        date=str(date)
        y,m,d=date.split("-")
        ws[char+'9'].value = f'{d}\n{m}\n{y[-2::]}\n{time}'
        
        roll=[]
        for col in range(10,n+11):
            roll.append(str(ws['B'+str(col)].value)[-3::])
        
        enroll=[str(i).zfill(3) for i in roll_list.split(",")]

        for col in range(10,n+10):
            if str(roll[col-10]) in enroll :
                object.add_frame(col,ws,"light green","P")
                ws[char+str(col)] =  'P'
            else :
                ws[char+str(col)] =  'A'
                object.add_frame(col,ws,"red","A")
            ws['W'+str(col)]=f'=COUNTIF(D{str(col)}:V{str(col)},"P")'
            ws['AG'+str(col)]=f'=SUM(W{str(col)},AF{str(col)})'
        
        ws[char+str(n+11)]=f'=COUNTIF({char}10:{char}{n-1},"P")' 
        ws['W'+str(n+11)]=f'=SUM(W10:W{n-1})'
        ws['AG'+str(n+11)]=f'=SUM(W10:W{n-1})'
        
        def savefun() :
            wb.save(f"workbooks\{user}\{Excel}")
            Thread(target=Drive_uplode.upload_files , args=(user,Excel)).start()
            messagebox.showinfo("save","Entry save sucessfully")
        
        object.extend_end(backfun,savefun,n)
        
        if char=="V":
            messagebox.showerror("","Sheet is full! Create new sheet")
    except Exception as e:
        messagebox.showerror("An Error occure",f"Entry not saved\n{e}")
        backfun()
        return

def enterpract(backfun,object,user,Excel,date,time,roll_list):
    try:      
        try:
            wb=load_workbook(f"workbooks\{user}\{Excel}")
        except:
            messagebox.showerror("Error","Workbook not found")
            return
        if len(wb.sheetnames)==1:
            messagebox.showerror("","Create new sheet")
            return
        
        with open("users/sheets.txt","r") as f :
            for name in f.readlines():
                if user+" "+Excel in name:
                    b=name.split(" ")
                    sheet=b[-2]
                    ws=wb[sheet]
                    break
            else:
                with open("users/sheets.txt","a") as f1 :
                    ws=wb.active
                    w=user+" "+Excel+" "+ws.title+" \n"
                    f1.write(w)
        
        n=numberofentry(ws)

        for row in range(24,32):
            char = get_column_letter(row)
            if ws[char+'9'].value == None  :
                break
        if char=="AE":
            messagebox.showerror("","Sheet is full! Create new sheet")
            return
        object.extend_start()
        date=str(date)
        y,m,d=date.split("-")
        ws[char+'9'].value = f'{d}\n{m}\n{y[-2::]}\n{time}'

        roll=[]
        for col in range(10,n+11):
            roll.append(str(ws['B'+str(col)].value)[-3::])
        
        enroll=[str(i).zfill(3) for i in roll_list.split(",")]

        for col in range(10,n+10):
            if str(roll[col-10]) in enroll :
                object.add_frame(col,ws,"light green","P")
                ws[char+str(col)] =  'P'
            else :
                object.add_frame(col,ws,"red","A")
                ws[char+str(col)] =  'A'
            ws['AF'+str(col)]=f'=COUNTIF(X{str(col)}:AE{str(col)},"P")'
            ws['AG'+str(col)]=f'=SUM(W{str(col)},AF{str(col)})'

        ws[char+str(n+11)]=f'=COUNTIF({char}10:{char}{n-1},"P")' 
        ws['AF'+str(n+11)]=f'=SUM(AF10:AF{n-1})'
        ws['AG'+str(n+11)]=f'=SUM(W10:W{n-1})'

        def savefun() :
            wb.save(f"workbooks\{user}\{Excel}")
            Thread(target=Drive_uplode.upload_files , args=(user,Excel)).start()
            messagebox.showinfo("save","Entry save sucessfully")
        
        object.extend_end(backfun,savefun,n)
        
        if char=="AE":
            messagebox.showerror("","Sheet is full! Create new sheet")
    except:
        messagebox.showerror("An Error occure","Entry not saved")
        backfun()
        return

def select(backfun,object,user,sheet,date,time,roll_list,tp):
    roll_list=roll_list[:-1:]
    if time == "" or time == " " or time=="Time" :
        messagebox.showerror("Error","All Fields Are Required")
        return
    if tp == "Theory":
        entertheory(backfun,object,user,sheet,date,time,roll_list)
    elif tp=="Practicle":
        enterpract(backfun,object,user,sheet,date,time,roll_list)
    else:
        messagebox.showerror("Error","Invalid choice")

