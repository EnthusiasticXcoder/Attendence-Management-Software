import csv
import os
import shutil

from openpyxl import load_workbook , Workbook
from openpyxl.utils import get_column_letter
from xlcalculator import ModelCompiler
from xlcalculator import Evaluator

from utilities.exception import (
    UnableToGetWorksheetException,
    CreateNewWorksheetException,
    UnableToCreateWorksheetException
)
from utilities.constants import WORKBOOKPATH, WORKSHEET
from utilities.constants.routes import SHEETDATA


class WorkBookService :
    __instance = None

    def __new__(cls):
        if cls.__instance is None :
            return super().__new__(cls)
        return cls.__instance

    def __init__(self) -> None:
        WorkBookService.__instance = self

    def addWorkbook(self, dirname: str , filesrc: str):
        ''' A Method to Add Excel Workbook File To the application '''
        filename=os.path.basename(filesrc)
        filedst =os.path.join(os.getcwd(), dirname , filename)
        shutil.copyfile(filesrc, filedst)   
    
    def CreateNewWorksheet(self, wbPath: str, title: str):
        try : 
            workbook = load_workbook(filename= wbPath)
        except Exception :
            raise FileNotFoundError
        
        worksheet=workbook.active
        NewWorksheet=workbook.copy_worksheet(worksheet)
        NewWorksheet.title= title
        
        dataRow = []
        try :
            with open( SHEETDATA ,mode='r') as file :
                reader = csv.DictReader(file,fieldnames=[WORKBOOKPATH, WORKSHEET])
                for data in reader :
                    if data[WORKBOOKPATH] == wbPath :
                        continue
                    dataRow.append(data)
            
            with open( SHEETDATA ,mode='w') as file :
                writer = csv.DictWriter(file,fieldnames=[WORKBOOKPATH, WORKSHEET])
                writer.writerows(dataRow)
                writer.writerow({WORKBOOKPATH : wbPath,
                                WORKSHEET : title})
            workbook.save(wbPath)
        except Exception :
            raise UnableToCreateWorksheetException

    def EnterAttendence(self, CSVRollno: str, DateTime: str, isTheory: bool, wbPath: str  ):
        ''' A Method to Enter Attendence in the worksheet '''
        try : 
            workbook = load_workbook(filename= wbPath)
        except Exception :
            raise FileNotFoundError
        
        if workbook.sheetnames.__len__() == 1 :
            raise CreateNewWorksheetException

        worksheet = self._getWorkSheet(workbook=workbook, wbpath= wbPath)

        numberofentry = self.getNumberofEntry(worksheet)
        letter = self._getCharacter(worksheet=worksheet,isTheory=isTheory)

        worksheet[f'{letter}9'].value = DateTime

        EnrollmentList=[]
        for column in range(10,numberofentry+10):
            EnrollmentList.append(str(worksheet[f'B{column}'].value))
        
        Roll_list=[number.zfill(3).capitalize() for rollno in CSVRollno.split(",") for number in str(rollno).strip().split()]

        temp = []
        for column,roll in enumerate(EnrollmentList) :    
            for cheack in Roll_list:
                if roll.endswith(cheack):
                    worksheet[f'{letter}{column+10}'] = 'P'
                    temp.append('P')
                    Roll_list.remove(cheack)
                    break
            else :
                worksheet[f'{letter}{column+10}'] = 'A'
                temp.append('A')
        
        return EnrollmentList,worksheet[f'C10:C{numberofentry+10}'], temp, lambda: workbook.save(wbPath)
        
    def get_worksheet_from_path(self, wbpath: str):
        ''' get list of all worksheet instance in the workbook '''
        try : 
            workbook = load_workbook(filename= wbpath)
        except Exception :
            raise FileNotFoundError
        
        sheetobject = [ workbook[sheet] for sheet in workbook.get_sheet_names()]
        return sheetobject
    
    def get_evaluator(self, wbpath: str):
        compiler = ModelCompiler()
        new_model = compiler.read_and_parse_archive(wbpath)
        evaluator = Evaluator(new_model)
        return evaluator
        
    def _getCharacter(self, worksheet , isTheory):
        if isTheory :
            for row in range(4,23):
                char = get_column_letter(row)
                if worksheet[f'{char}9'].value == None  :
                    break
            else :
                raise CreateNewWorksheetException
        else :
            for row in range(24,32):
                char = get_column_letter(row)
                if worksheet[f'{char}9'].value == None : 
                    break
            else :
                raise CreateNewWorksheetException
        return char
    
    def getNumberofEntry(self, worksheet):
        n=0
        while True:
            if worksheet[f'A{n+10}'].value == None : 
                break
            n=n+1
        return n
    
    def _getWorkSheet(self, workbook: Workbook, wbpath: str) :
        ''' A Method To return worksheet '''
        try :
            with open( SHEETDATA ,mode='r') as file :
                reader = csv.DictReader(file,fieldnames=[WORKBOOKPATH, WORKSHEET])
                for data in reader :
                    if data[WORKBOOKPATH] == wbpath :
                        return workbook[data[WORKSHEET]]
                else :
                    with open( SHEETDATA ,mode='a') as file :
                        writer = csv.DictWriter(file,fieldnames=[WORKBOOKPATH, WORKSHEET])
                        worksheet = workbook.active
                        writer.writerow({WORKBOOKPATH : wbpath,
                                        WORKSHEET : worksheet.title})
                        return worksheet
        except Exception :
            raise UnableToGetWorksheetException

